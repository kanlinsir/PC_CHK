
# -*- coding: UTF-8
import os
import socket
import subprocess
import json
import winreg
import platform
from openpyxl import Workbook

def main():
    computer_name = platform.node()  # 等同於 Windows 的電腦名稱
    local_ip = get_local_ip_address()  # 取得第一個 IPv4 位址

    # 建立 Excel 工作簿
    wb = Workbook()
    
    # 1. 系統資訊
    ws_system = wb.active
    ws_system.title = "系統資訊"
    system_info = get_system_info()
    write_dict_list_to_sheet(ws_system, system_info, ["項目", "內容"])
    
    # 2. Windows Defender 資訊
    ws_defender = wb.create_sheet("Windows Defender")
    defender_info = get_defender_info()
    write_dict_list_to_sheet(ws_defender, defender_info, [
        "AMProductVersion", 
        "AMServiceVersion", 
        "AntispywareSignatureVersion", 
        "AntivirusSignatureVersion"
    ])

    # 3. 已安裝更新 (QuickFixEngineering)
    ws_updates = wb.create_sheet("已安裝更新")
    updates = get_installed_updates()
    write_dict_list_to_sheet(ws_updates, updates, ["HotFixID", "Description", "InstalledOn"])

    # 4. 已安裝程式 (從登錄讀取)
    ws_programs = wb.create_sheet("已安裝程式")
    programs = get_installed_programs()
    write_dict_list_to_sheet(ws_programs, programs, ["名稱", "版本", "發行者"])

    # 5. 使用者帳號 (Win32_UserAccount)
    ws_users = wb.create_sheet("使用者帳號")
    user_accounts = get_local_user_accounts()
    write_dict_list_to_sheet(ws_users, user_accounts, ["帳號名稱", "網域", "描述", "是否啟用"])

    # 6. 密碼原則 (net accounts)
    ws_policy = wb.create_sheet("密碼原則")
    password_policy = get_password_policy()
    write_dict_list_to_sheet(ws_policy, password_policy, ["設定", "值"])

    # 7. 網路設定 (列出所有 IP、子網遮罩及 DNS server)
    ws_network = wb.create_sheet("網路設定")
    network_settings = get_network_settings()
    write_dict_list_to_sheet(ws_network, network_settings, ["介面名稱", "IP位址", "子網遮罩", "DNS server"])

    # 儲存 Excel，檔名以「電腦名稱_第一個區網IP.xlsx」命名
    filename = f"{computer_name}_{local_ip}.xlsx"
    if os.path.exists(filename):
        response = input(f"檔案 {filename} 已存在，是否覆蓋? (Y/N): ")
        if response.strip().lower() != 'y':
            print("取消覆蓋，程式終止。")
            return
    wb.save(filename)
    print(f"系統資訊已匯出至 {filename}")

#------------------------------------------------------------------------------
# 1. 取得系統資訊 (電腦名稱、IP、OS版本)
#------------------------------------------------------------------------------

def get_system_info():
    data = []
    data.append({"項目": "電腦名稱", "內容": platform.node()})
    data.append({"項目": "區網 IP", "內容": get_local_ip_address()})
    data.append({"項目": "Windows 版本", "內容": platform.platform()})
    return data

def get_local_ip_address():
    """只抓第一個 IPv4（排除 127.*）"""
    host_name = socket.gethostname()
    for addr_info in socket.getaddrinfo(host_name, None):
        ip = addr_info[4][0]
        if ":" not in ip and not ip.startswith("127."):
            return ip
    return "未知"

def get_external_ip_address():
    """若需要對外 IP，可使用此函式 (需連上外網)"""
    try:
        result = subprocess.run(
            ["powershell", "-Command", "Invoke-RestMethod https://api.ipify.org"],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            return result.stdout.strip()
    except:
        pass
    return "未知"

#------------------------------------------------------------------------------
# 2. 取得 Windows Defender 資訊 (以 JSON 格式)
#------------------------------------------------------------------------------

def get_defender_info():
    """
    以 PowerShell 執行 Get-MpComputerStatus，擷取四個欄位後轉成 JSON。
    回傳 List[Dict] 方便寫入 Excel。
    """
    ps_cmd = r'''
Get-MpComputerStatus |
Select-Object AMProductVersion,AMServiceVersion,AntispywareSignatureVersion,AntivirusSignatureVersion |
ConvertTo-Json
'''
    result_json = run_powershell(ps_cmd)
    try:
        data = json.loads(result_json)
        if isinstance(data, dict):
            data = [data]
        return data
    except:
        return [{"AMProductVersion": "無法取得",
                 "AMServiceVersion": "",
                 "AntispywareSignatureVersion": "",
                 "AntivirusSignatureVersion": ""}]

#------------------------------------------------------------------------------
# 3. 取得已安裝更新 (QuickFixEngineering)
#------------------------------------------------------------------------------

def get_installed_updates():
    """
    透過 PowerShell 取得 Win32_QuickFixEngineering 的 HotFixID、Description、InstalledOn，
    轉成 JSON 再解析。
    """
    ps_cmd = r'''
Get-WmiObject Win32_QuickFixEngineering |
Select-Object HotFixID,Description,InstalledOn |
ConvertTo-Json
'''
    result_json = run_powershell(ps_cmd)
    try:
        data = json.loads(result_json)
        if isinstance(data, dict):
            data = [data]
        return data
    except:
        return []

#------------------------------------------------------------------------------
# 4. 取得已安裝程式 (從登錄讀取)
#------------------------------------------------------------------------------

def get_installed_programs():
    r"""
    從以下登錄路徑讀取已安裝程式資訊：
      HKLM\SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall
      HKLM\SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall
    以及 HKCU 的對應路徑。
    回傳 List[Dict]，每個 Dict 包含：名稱、版本、發行者。
    """
    data = []
    registry_paths = [
        r"SOFTWARE\Microsoft\Windows\CurrentVersion\Uninstall",
        r"SOFTWARE\WOW6432Node\Microsoft\Windows\CurrentVersion\Uninstall"
    ]
    hives = [
        (winreg.HKEY_LOCAL_MACHINE, "HKLM"),
        (winreg.HKEY_CURRENT_USER, "HKCU")
    ]
    for hive, hive_name in hives:
        for path in registry_paths:
            try:
                with winreg.OpenKey(hive, path) as reg_key:
                    for i in range(winreg.QueryInfoKey(reg_key)[0]):
                        subkey_name = winreg.EnumKey(reg_key, i)
                        try:
                            with winreg.OpenKey(reg_key, subkey_name) as subkey:
                                display_name = get_reg_value(subkey, "DisplayName")
                                if not display_name:
                                    continue
                                display_version = get_reg_value(subkey, "DisplayVersion")
                                publisher = get_reg_value(subkey, "Publisher")
                                data.append({
                                    "名稱": display_name,
                                    "版本": display_version if display_version else "",
                                    "發行者": publisher if publisher else ""
                                })
                        except:
                            pass
            except:
                pass
    return data

def get_reg_value(key, value_name):
    try:
        return winreg.QueryValueEx(key, value_name)[0]
    except:
        return None

#------------------------------------------------------------------------------
# 5. 取得本機使用者帳號 (Win32_UserAccount, LocalAccount=True)
#------------------------------------------------------------------------------

def get_local_user_accounts():
    """
    透過 PowerShell 呼叫 Get-WmiObject Win32_UserAccount -Filter "LocalAccount=True"
    並轉成 JSON，再解析回傳 List[Dict]。
    """
    ps_cmd = r'''
Get-WmiObject Win32_UserAccount -Filter "LocalAccount=True" |
Select-Object Name,Domain,Description,Disabled |
ConvertTo-Json
'''
    result_json = run_powershell(ps_cmd)
    try:
        data = json.loads(result_json)
        if isinstance(data, dict):
            data = [data]
        result = []
        for item in data:
            disabled = item.get("Disabled", False)
            result.append({
                "帳號名稱": item.get("Name", ""),
                "網域": item.get("Domain", ""),
                "描述": item.get("Description", ""),
                "是否啟用": "停用" if disabled else "啟用"
            })
        return result
    except:
        return []

#------------------------------------------------------------------------------
# 6. 取得密碼原則 (net accounts)
#------------------------------------------------------------------------------

def get_password_policy():
    """
    執行 net accounts 指令，將文字輸出解析為 List[Dict]。
    """
    raw_output = run_powershell("net accounts")
    lines = raw_output.splitlines()
    data = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        parts = line.split(":", 1)
        if len(parts) == 2:
            key = parts[0].strip()
            val = parts[1].strip()
            data.append({"設定": key, "值": val})
    return data

#------------------------------------------------------------------------------
# 7. 取得網路設定 (IP、子網遮罩及 DNS server)
#------------------------------------------------------------------------------

def get_network_settings():
    """
    利用 PowerShell 及 WMI 查詢 IPEnabled=True 的網路介面，
    取得 Description、IPAddress、IPSubnet 及 DNSServerSearchOrder，
    回傳 List[Dict]，每筆資料包含：介面名稱、IP位址、子網遮罩、DNS server。
    """
    ps_cmd = r'''
Get-WmiObject Win32_NetworkAdapterConfiguration -Filter "IPEnabled=True" |
Select-Object Description, IPAddress, IPSubnet, DNSServerSearchOrder |
ConvertTo-Json
'''
    result_json = run_powershell(ps_cmd)
    try:
        data = json.loads(result_json)
        if isinstance(data, dict):
            data = [data]
        network_settings = []
        for item in data:
            interface = item.get("Description", "")
            ip_list = item.get("IPAddress", [])
            subnet_list = item.get("IPSubnet", [])
            dns_list = item.get("DNSServerSearchOrder", [])
            ip_str = ", ".join(ip_list) if isinstance(ip_list, list) else str(ip_list)
            subnet_str = ", ".join(subnet_list) if isinstance(subnet_list, list) else str(subnet_list)
            dns_str = ", ".join(dns_list) if (isinstance(dns_list, list) and dns_list) else (str(dns_list) if dns_list else "")
            network_settings.append({
                "介面名稱": interface,
                "IP位址": ip_str,
                "子網遮罩": subnet_str,
                "DNS server": dns_str
            })
        return network_settings
    except Exception as e:
        print("Error parsing network settings:", e)
        return [{"介面名稱": "無", "IP位址": "無", "子網遮罩": "無", "DNS server": "無法取得"}]

#------------------------------------------------------------------------------
# 8. 呼叫 PowerShell
#------------------------------------------------------------------------------

def run_powershell(command):
    """
    呼叫 PowerShell 指令，回傳標準輸出 (str)。
    使用 subprocess.run 並捕捉 output。
    """
    proc = subprocess.run(["powershell", "-Command", command],
                          capture_output=True, text=True)
    return proc.stdout

#------------------------------------------------------------------------------
# 9. 將 List[Dict] 寫入 Excel 工作表
#------------------------------------------------------------------------------

def write_dict_list_to_sheet(ws, dict_list, field_order=None):
    """
    將 List[Dict] 的資料寫入 openpyxl Worksheet。
    若某欄位值為 dict 且含有 "DateTime" 鍵，則取該鍵的值；否則以 str() 轉換。
    """
    if not dict_list:
        ws.append(["無資料"])
        return

    if not field_order:
        field_order = list(dict_list[0].keys())

    # 寫入表頭
    ws.append(field_order)

    # 寫入每一筆資料
    for item in dict_list:
        row = []
        for field in field_order:
            value = item.get(field, "")
            if isinstance(value, dict):
                if "DateTime" in value:
                    value = value["DateTime"]
                else:
                    value = str(value)
            row.append(value)
        ws.append(row)

#------------------------------------------------------------------------------

if __name__ == "__main__":
    main()
